#!/usr/bin/env python3
"""
Analyze Pocket export tags and create binary columns for each unique tag.
Filters for 'unread' status in column E and analyzes tags in column N.
"""

import openpyxl
from collections import Counter
import re

# Read the Excel file
wb = openpyxl.load_workbook('/Users/petermaxfield/Library/CloudStorage/GoogleDrive-pmaxfield@gmail.com/My Drive/Files/Files_2025/Pocket_export/pocket export 20250522.xlsx')
ws = wb.active

# Get headers
headers = [cell.value for cell in ws[1]]
print(f"Column E: {headers[4]}")
print(f"Column N: {headers[13]}")
print(f"\nTotal rows: {ws.max_row - 1}")

# Find unread items and collect all tags
all_tags = []
unread_rows = []

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    status = row[4].value  # Column E (0-indexed as 4)
    tags = row[13].value   # Column N (0-indexed as 13)
    
    if status and str(status).lower() == 'unread':
        unread_rows.append(row_idx)
        if tags:
            # Split tags by comma first, then by pipe delimiter
            comma_split = [tag.strip() for tag in str(tags).split(',')]
            tag_list = []
            for tag in comma_split:
                # Split by pipe and clean each part
                pipe_split = [t.strip().lower() for t in tag.split('|') if t.strip()]
                tag_list.extend(pipe_split)
            all_tags.extend(tag_list)

print(f"Unread rows: {len(unread_rows)}")
print(f"\nTotal tags found: {len(all_tags)}")

# Count unique tags
tag_counts = Counter(all_tags)
print(f"Unique tags: {len(tag_counts)}")

# Sort tags by frequency
sorted_tags = sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)

print("\n" + "="*80)
print("TAG FREQUENCY ANALYSIS")
print("="*80)
for tag, count in sorted_tags:
    print(f"{tag:40} : {count:4} occurrences")

# Suggest tags that should be combined
print("\n" + "="*80)
print("SUGGESTED TAG COMBINATIONS")
print("="*80)

suggestions = []
tags_list = list(tag_counts.keys())

for i, tag1 in enumerate(tags_list):
    for tag2 in tags_list[i+1:]:
        # Check for similar tags
        if tag1 in tag2 or tag2 in tag1:
            suggestions.append((tag1, tag2, tag_counts[tag1], tag_counts[tag2]))
        # Check for common variations
        elif tag1.replace('-', ' ') == tag2.replace('-', ' '):
            suggestions.append((tag1, tag2, tag_counts[tag1], tag_counts[tag2]))
        elif tag1.replace('_', ' ') == tag2.replace('_', ' '):
            suggestions.append((tag1, tag2, tag_counts[tag1], tag_counts[tag2]))

if suggestions:
    for tag1, tag2, count1, count2 in suggestions:
        print(f"  • '{tag1}' ({count1}) ↔ '{tag2}' ({count2})")
else:
    print("No obvious tag combinations found.")

# Create new workbook with tag columns
print("\n" + "="*80)
print("CREATING NEW WORKBOOK WITH TAG COLUMNS")
print("="*80)

# Create output workbook
output_wb = openpyxl.Workbook()
output_ws = output_wb.active

# Write original headers
for col_idx, header in enumerate(headers, start=1):
    output_ws.cell(row=1, column=col_idx, value=header)

# Add new tag column headers
unique_tags = sorted(tag_counts.keys())
start_col = len(headers) + 1
for idx, tag in enumerate(unique_tags):
    output_ws.cell(row=1, column=start_col + idx, value=f"tag_{tag}")

print(f"Added {len(unique_tags)} new tag columns")

# Write data for unread rows only
output_row = 2
for row_idx in unread_rows:
    row = list(ws[row_idx])
    
    # Write original data
    for col_idx, cell in enumerate(row, start=1):
        output_ws.cell(row=output_row, column=col_idx, value=cell.value)
    
    # Get tags for this row
    tags_cell = row[13].value
    row_tags = []
    if tags_cell:
        # Split tags by comma first, then by pipe delimiter
        comma_split = [tag.strip() for tag in str(tags_cell).split(',')]
        for tag in comma_split:
            # Split by pipe and clean each part
            pipe_split = [t.strip().lower() for t in tag.split('|') if t.strip()]
            row_tags.extend(pipe_split)
    
    # Mark tag columns
    for idx, tag in enumerate(unique_tags):
        tag_col = start_col + idx
        output_ws.cell(row=output_row, column=tag_col, value=1 if tag in row_tags else 0)
    
    output_row += 1

# Save output file
output_path = '/Users/petermaxfield/Library/CloudStorage/GoogleDrive-pmaxfield@gmail.com/My Drive/Files/Files_2025/Pocket_export/pocket_export_with_tags_20251019.xlsx'
output_wb.save(output_path)

print(f"\nOutput saved to: {output_path}")
print(f"Rows in output: {output_row - 2}")
print(f"Columns in output: {len(headers) + len(unique_tags)}")

