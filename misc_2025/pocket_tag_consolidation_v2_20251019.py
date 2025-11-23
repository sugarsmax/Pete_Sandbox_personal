#!/usr/bin/env python3
"""
Consolidate and clean up tags in the Pocket export file - Version 2.
Applies additional tag mappings based on review.
"""

import openpyxl
from collections import defaultdict

# Define tag consolidation mappings
TAG_MAPPINGS = {
    # From previous consolidation (keeping these)
    'periodicals': 'periodical',
    'sports': 'sport',
    'materials': 'material',
    'graphics': 'graphic',
    'graphs': 'graph',
    'charts': 'chart',
    'drinks': 'drink',
    'shows': 'show',
    'figures': 'figure',
    'crafts': 'craft',
    'courses': 'course',
    'kids': 'kid',
    'fires': 'fire',
    
    # Manufacturing
    'cam': 'mfg',
    'cnc': 'mfg',
    
    # UX and IA
    'uxstrategy': 'ux',
    'uxr': 'ux',
    'ia': 'ux',
    
    # Simulation
    'ansys': 'sim',
    'fea': 'sim',
    'simulation': 'sim',
    
    # Generative Design
    'gd': 'generative',
    'dreamcatcher': 'generative',
    
    # Autodesk/CAD companies
    'adsk': 'autodesk',
    'develop3d': 'autodesk',
    'fusion': 'autodesk',
    'fusion360': 'autodesk',
    'mae': 'autodesk',
    'onshape': 'autodesk',
    'ptc': 'autodesk',
    'au': 'autodesk',  # NEW: abbreviation
    'forums': 'autodesk',  # NEW: Autodesk forums
    
    # AI
    'aitools': 'ai',
    'openai': 'ai',
    'chatgpt': 'ai',
    'ml': 'ai',
    'deepfake': 'ai',
    'al': 'ai',
    
    # Art
    'artwork': 'art',
    'durer': 'art',
    
    # Automotive
    'audi': 'automotive',
    'car': 'automotive',
    'rivian': 'automotive',
    'ev': 'automotive',
    'car_ev': 'automotive',
    'tesla': 'automotive',
    'etron': 'automotive',
    'formula-e': 'automotive',
    'f1': 'automotive',
    'racing': 'automotive',
    
    # Visualization
    'viz': 'visualization',
    'vis': 'visualization',
    'graph': 'visualization',  # NEW: graph → visualization
    'rgraphing': 'visualization',
    'datavis': 'visualization',
    
    # Music services
    'last': 'last.fm',
    'soundcloud': 'music',
    'piano': 'music',
    'sheetmusic': 'music',
    'music_other': 'music',
    'music converters': 'music',
    'music_movies': 'music',
    'qotsa': 'music',
    'death from above': 'music',
    'theheavyeyes': 'music',
    'mantra': 'music',
    'reznor': 'music',
    'mp3': 'music',
    'audio': 'music',
    'foley': 'music',
    'sound effects': 'music',
    'sounds': 'music',
    
    # Mental health
    'anxiety': 'mentalhealth',
    'depression': 'mentalhealth',
    'selfcare': 'mentalhealth',
    'mindfulness': 'mentalhealth',
    'burnout': 'mentalhealth',
    'health': 'mentalhealth',
    'counciling': 'mentalhealth',
    'coping': 'mentalhealth',
    'pain': 'mentalhealth',
    'adhd': 'mentalhealth',
    'meditate': 'mentalhealth',
    
    # Typos and variants
    'quantifiedself': 'quantifiedself',
    'quatifiedself': 'quantifiedself',
    'personalityh': 'personality',
    'xds': 'xd',
    'loo': 'looker',
    'opt': 'optimization',
    
    # Related concepts
    'moldflow': 'mold',
    'molding': 'mold',
    'management': 'timemanagement',
    'parkingkitty': 'parking',
    'stereography': 'stereo',
    'photometry': 'photo',
    'spherical': 'photo',
    'insta360': 'camera',
    'gopro': 'camera',
    'dashcam': 'camera',
    'snapcam': 'camera',
    'stabilization': 'camera',
    
    # Portland/PDX
    'portland': 'pdx',
    'parkingkitty': 'pdx',
    'trimet': 'pdx',  # NEW: Portland transit
    'goearly': 'commute',  # NEW: commute app
    
    # House/remodel
    'house remodel 2021': 'remodel',
    'carpet': 'remodel',
    'paint': 'remodel',
    'kitchen': 'remodel',
    'couch': 'remodel',
    'heater': 'remodel',
    'furniture': 'remodel',
    'house': 'remodel',  # NEW: house → remodel
    
    # Tech companies
    'google': 'tech',
    'apple': 'tech',
    'microsoft': 'tech',
    'nvidia': 'tech',
    'aws': 'tech',
    'techx': 'tech',  # NEW: techx → tech
    
    # VR/AR - NEW: AR is augmented reality, align with VR
    'ar': 'vr',
    'oculus': 'vr',
    'quest': 'vr',
    'vr quest': 'vr',
    'hololens': 'vr',
    'xr': 'vr',
    
    # Lego
    'toy': 'lego',
    'transformers': 'lego',
    'walle': 'lego',
    'robot': 'lego',
    'arm': 'lego',
    
    # Office/work
    'msoffice': 'office',
    'workspace': 'office',
    'wfh': 'office',
    
    # Internet/network
    'internet': 'net',
    'speedtest': 'net',
    
    # Education
    'cornell': 'education',
    'bcs': 'education',
    'school': 'education',
    'julianna': 'education',
    'concordacademy': 'education',
    'ca': 'education',  # NEW: Concord Academy
    '529': 'education',
    'homework': 'education',
    
    # Games
    'pcgames': 'game',
    'xbox': 'game',
    'playstation': 'game',
    'atari': 'game',
    'minecraft': 'game',
    'blackandwhite': 'game',
    'arcade': 'game',  # NEW: arcade → game (not CAD)
    
    # Periodicals
    'ieee': 'periodical',
    'mittechreview': 'periodical',
    'nyt': 'periodical',
    'newyorker': 'periodical',
    
    # Social media
    'reddit': 'social',
    'tiktok': 'social',
    'snapchat': 'social',
    'instagram': 'social',
    
    # Family
    'jamie': 'family',
    'stow': 'family',
    'parenting': 'family',
    'dad': 'family',
    
    # Renewable energy
    'solar': 'renewable',
    'powerstation': 'renewable',
    'mirror': 'renewable',
    'sustainability': 'renewable',  # NEW: sustainability → renewable
    
    # Food
    'recipe': 'food',
    'cooking': 'food',
    'pretzel': 'food',
    'drinks': 'food',
    
    # Shopping
    'ikea': 'shopping',
    'amazon': 'shopping',
    'craigslist': 'shopping',
    'marketplace': 'shopping',
    'catalog': 'shopping',
    'gear': 'shopping',
    
    # Earthquake prep / Emergency
    'quake': 'emergency',  # NEW: consolidated to emergency
    'cascadia': 'emergency',
    'earthquake': 'emergency',
    'prep': 'emergency',
    'p4p': 'emergency',  # NEW: Parents for Preparedness
    'emergency': 'emergency',
    
    # Space/Astronomy - NEW
    'mars': 'space',
    'nasa': 'space',
    'stars': 'space',
    'planetearth': 'space',
    
    # Design tools - NEW
    'figma': 'designtools',
    'mural': 'designtools',
    'luma': 'designtools',
    'tools': 'designtools',
    
    # CAD - NEW
    'tinkercad': 'cad',
    
    # Programming - NEW
    'programming': 'coding',
    'python': 'coding',
    'javascript': 'coding',
    
    # Hardware - NEW
    'keyboard': 'hardware',
    'mac': 'hardware',
    
    # Data/Analytics - NEW
    'data': 'analytics',
    'voc': 'analytics',
    
    # Geography/Mapping - NEW
    'map': 'geography',
    'gis': 'geography',
    
    # Research/Learning - NEW
    'research': 'learning',
    'hci': 'learning',
    
    # Career - NEW
    'speaking': 'career',
    
    # Misc consolidations
    'biography': 'book',
    'kindle': 'book',
    'documentary': 'video',
    'netflix': 'video',
    'movie': 'video',
    'tv': 'video',
    'machinedesign': 'design',  # NEW
}

print("Loading workbook...")
wb = openpyxl.load_workbook('/Users/petermaxfield/Library/CloudStorage/GoogleDrive-pmaxfield@gmail.com/My Drive/Files/Files_2025/Pocket_export/pocket_export_consolidated_20251019.xlsx')
ws = wb.active

# Get headers
headers = [cell.value for cell in ws[1]]
print(f"Total columns: {len(headers)}")

# Find where tag columns start (after original columns)
original_col_count = 21  # Based on previous output
tag_start_col = original_col_count + 1

# Get all tag column names
tag_columns = headers[original_col_count:]
print(f"Tag columns: {len(tag_columns)}")

# Create mapping of old tag names to new consolidated names
tag_name_map = {}
for tag in tag_columns:
    # Remove 'tag_' prefix if present
    clean_tag = tag.replace('tag_', '') if tag.startswith('tag_') else tag
    
    # Apply consolidation mapping
    if clean_tag in TAG_MAPPINGS:
        new_tag = TAG_MAPPINGS[clean_tag]
    else:
        new_tag = clean_tag
    
    tag_name_map[tag] = new_tag

# Count unique consolidated tags
unique_consolidated_tags = sorted(set(tag_name_map.values()))
print(f"\nConsolidated to {len(unique_consolidated_tags)} unique tags")

# Show consolidation summary
consolidation_summary = defaultdict(list)
for old_tag, new_tag in tag_name_map.items():
    old_clean = old_tag.replace('tag_', '')
    if old_clean != new_tag:
        consolidation_summary[new_tag].append(old_clean)

print("\n" + "="*80)
print("CONSOLIDATION SUMMARY (NEW MAPPINGS ONLY)")
print("="*80)
new_mappings = ['vr', 'tech', 'education', 'game', 'renewable', 'emergency', 'space', 
                'designtools', 'cad', 'coding', 'hardware', 'analytics', 'geography', 
                'learning', 'career', 'remodel', 'pdx', 'commute', 'design', 'autodesk', 
                'visualization']
for new_tag in sorted(consolidation_summary.keys()):
    if new_tag in new_mappings:
        old_tags = consolidation_summary[new_tag]
        print(f"{new_tag:30} ← {', '.join(sorted(old_tags))}")

# Create new workbook with consolidated tags
print("\n" + "="*80)
print("CREATING CONSOLIDATED WORKBOOK V2")
print("="*80)

output_wb = openpyxl.Workbook()
output_ws = output_wb.active

# Write original column headers
for col_idx in range(1, original_col_count + 1):
    output_ws.cell(row=1, column=col_idx, value=headers[col_idx - 1])

# Write consolidated tag column headers
for idx, tag in enumerate(unique_consolidated_tags, start=1):
    output_ws.cell(row=1, column=original_col_count + idx, value=f"tag_{tag}")

print(f"New column count: {original_col_count + len(unique_consolidated_tags)}")

# Process each data row
print("Processing rows...")
for row_idx in range(2, ws.max_row + 1):
    if row_idx % 100 == 0:
        print(f"  Processing row {row_idx}/{ws.max_row}")
    
    row = list(ws[row_idx])
    
    # Copy original columns
    for col_idx in range(original_col_count):
        output_ws.cell(row=row_idx, column=col_idx + 1, value=row[col_idx].value)
    
    # Consolidate tag columns
    consolidated_tags = defaultdict(int)
    for col_idx in range(original_col_count, len(headers)):
        old_tag = headers[col_idx]
        new_tag = tag_name_map[old_tag]
        value = row[col_idx].value
        
        if value == 1:
            consolidated_tags[new_tag] = 1
    
    # Write consolidated tag values
    for idx, tag in enumerate(unique_consolidated_tags, start=1):
        output_ws.cell(row=row_idx, column=original_col_count + idx, 
                      value=consolidated_tags.get(tag, 0))

# Save output
output_path = '/Users/petermaxfield/Library/CloudStorage/GoogleDrive-pmaxfield@gmail.com/My Drive/Files/Files_2025/Pocket_export/pocket_export_consolidated_v2_20251019.xlsx'
output_wb.save(output_path)

print(f"\nOutput saved to: {output_path}")
print(f"Original tags (v1): {len(tag_columns)}")
print(f"Consolidated tags (v2): {len(unique_consolidated_tags)}")
print(f"Additional reduction: {len(tag_columns) - len(unique_consolidated_tags)} tags")
print(f"Total reduction from original 485: {485 - len(unique_consolidated_tags)} tags ({100*(485 - len(unique_consolidated_tags))/485:.1f}%)")

